from datetime import datetime
from django.utils import timezone
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import razorpay
from django.contrib import messages
from django.db.models import Sum
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl.styles import Font
import openpyxl
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils.text import slugify
import logging
from django.utils.dateparse import parse_datetime
from .prediction import predict_event_success
from django.db import transaction
from io import BytesIO
from django.views.decorators.http import require_POST
import secrets
import re
from django.core.exceptions import ValidationError
from django.core.files.images import get_image_dimensions
import mimetypes
from django.core.paginator import Paginator
from communities.models import ChatMessage
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import os
import pytz
import decimal

logger = logging.getLogger(__name__)

client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

from communities.models import Community, UserCommunity
from .models import Event, EventCategory, EventRegistration, Payment, VolunteerRegistration, EventGalleryItem


@login_required
def create_event(request, community_id):
    community = get_object_or_404(Community, id=community_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        category_id = request.POST.get('category')
        
        # Use parse_datetime instead of strptime
        start_datetime = parse_datetime(request.POST.get('start_datetime'))
        end_datetime = parse_datetime(request.POST.get('end_datetime'))
        
        # Make sure the datetimes are timezone-aware
        if start_datetime and not start_datetime.tzinfo:
            start_datetime = timezone.make_aware(start_datetime)
        if end_datetime and not end_datetime.tzinfo:
            end_datetime = timezone.make_aware(end_datetime)
        
        address = request.POST.get('address')
        location = request.POST.get('location')
        location_description = request.POST.get('location_description')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        registration_fee = request.POST.get('registration_fee')
        cover_image = request.FILES.get('cover_image')
        max_attendees = request.POST.get('max_attendees')
        
        # Handle pricing fields
        base_price = request.POST.get('base_price', 0)
        dynamic_pricing_enabled = request.POST.get('dynamic_pricing_enabled') == 'on'
        
        if dynamic_pricing_enabled:
            try:
                min_price = decimal.Decimal(request.POST.get('min_price', base_price))
                max_price = decimal.Decimal(request.POST.get('max_price', base_price))
                price_increment = decimal.Decimal(request.POST.get('price_increment_percentage', 10))
            except decimal.InvalidOperation:
                messages.error(request, "Please enter valid prices")
                return redirect('events:create_event', community_id=community_id)
        else:
            min_price = max_price = base_price
            price_increment = 10

        category = get_object_or_404(EventCategory, id=category_id)
        
        try:
            event = Event(
                name=name,
                description=description,
                category=category,
                start_datetime=start_datetime,
                end_datetime=end_datetime,
                address=address,
                location=location,
                location_description=location_description,
                latitude=latitude,
                longitude=longitude,
                base_price=base_price,
                dynamic_pricing_enabled=dynamic_pricing_enabled,
                min_price=min_price,
                max_price=max_price,
                price_increment_percentage=price_increment,
                organizer=request.user,
                community=community,
                cover_image=cover_image,
                max_attendees=max_attendees
            )
            event.full_clean()  # This will run validation
            event.save()
            return redirect('events:event_detail', event_id=event.id)
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('events:create_event', community_id=community_id)

    categories = EventCategory.objects.filter(status="enabled")
    return render(request, 'events/create_event.html', {'categories': categories, 'community': community})

@login_required
def update_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    if request.method == 'POST':
        # Get the new max_attendees value
        new_max_attendees = int(request.POST.get('max_attendees', 50))
        
        # Get current registration count
        current_registrations = EventRegistration.objects.filter(event=event).count()
        
        # Validate that new limit isn't less than current registrations
        if new_max_attendees < current_registrations:
            messages.error(request, f"Cannot set maximum attendees below current registration count ({current_registrations})")
            return redirect('events:update_event', event_id=event.id)
        
        # Parse datetime strings and make them timezone-aware
        start_datetime = parse_datetime(request.POST.get('start_datetime'))
        end_datetime = parse_datetime(request.POST.get('end_datetime'))
        
        if start_datetime and end_datetime:
            # Make datetime objects timezone-aware
            ist_timezone = pytz.timezone('Asia/Kolkata')
            start_datetime = ist_timezone.localize(start_datetime)
            end_datetime = ist_timezone.localize(end_datetime)
            
            # Handle pricing fields
            base_price = request.POST.get('base_price', 0)
            dynamic_pricing_enabled = request.POST.get('dynamic_pricing_enabled') == 'on'
            
            if dynamic_pricing_enabled:
                try:
                    min_price = decimal.Decimal(request.POST.get('min_price', base_price))
                    max_price = decimal.Decimal(request.POST.get('max_price', base_price))
                    price_increment = decimal.Decimal(request.POST.get('price_increment_percentage', 10))
                except decimal.InvalidOperation:
                    messages.error(request, "Please enter valid prices")
                    return redirect('events:update_event', event_id=event.id)
            else:
                min_price = max_price = base_price
                price_increment = 10

            # Update event fields
            event.name = request.POST.get('name')
            event.description = request.POST.get('description')
            event.category_id = request.POST.get('category')
            event.start_datetime = start_datetime
            event.end_datetime = end_datetime
            event.address = request.POST.get('address')
            event.location = request.POST.get('location')
            event.location_description = request.POST.get('location_description')
            event.latitude = request.POST.get('latitude')
            event.longitude = request.POST.get('longitude')
            event.max_attendees = new_max_attendees
            
            if 'cover_image' in request.FILES:
                event.cover_image = request.FILES['cover_image']
            
            event.base_price = base_price
            event.dynamic_pricing_enabled = dynamic_pricing_enabled
            event.min_price = min_price
            event.max_price = max_price
            event.price_increment_percentage = price_increment
            
            event.full_clean()  # This will run validation
            event.save()
            messages.success(request, 'Event updated successfully')
            return redirect('events:event_detail', event_id=event.id)
        else:
            messages.error(request, 'Invalid datetime format')
            return redirect('events:update_event', event_id=event.id)

    categories = EventCategory.objects.filter(status="enabled")
    return render(request, 'events/update_event.html', {
        'event': event,
        'categories': categories
    })

@login_required
def event_management(request, community_id):
    user = request.user
    community = get_object_or_404(Community, id=community_id)
    unread_messages_count = ChatMessage.objects.filter(
        receiver=request.user,
        is_read=False
    ).count()
    # Check if the user is a member of the community
    if not UserCommunity.objects.filter(user=user, community=community).exists():
        return HttpResponse('Unauthorized', status=403)
    
    events = Event.objects.filter(organizer=user, community=community)
    return render(request, 'events/event_management.html', {'events': events, 'community': community, 'unread_messages_count': unread_messages_count})


@csrf_exempt
@login_required
def register_for_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Use dynamic price instead of fixed registration_fee
    current_price = event.current_price
    
    # Check if event has reached maximum capacity
    current_registrations = EventRegistration.objects.filter(event=event).count()
    if current_registrations >= event.max_attendees:
        return JsonResponse({
            'status': 'error',
            'message': 'Sorry, this event has reached its maximum capacity'
        }, status=400)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')

        if action == 'register_free':
            # Create event registration
            registration = EventRegistration.objects.create(
                user=request.user,
                event=event,
                status='confirmed'
            )
            registration.generate_qr_code()
            
            # Send confirmation email with QR code
            send_registration_confirmation_email(request.user, event, registration)
            
            return JsonResponse({'status': 'success', 'message': 'Registration successful'})

        elif action == 'create_order':
            # Create Razorpay Order
            order_amount = int(current_price * 100)  # Convert to paise
            order_currency = 'INR'
            order_receipt = f'order_rcptid_{event.id}_{request.user.id}'
            razorpay_order = client.order.create(dict(
                amount=order_amount,
                currency=order_currency,
                receipt=order_receipt
            ))

            # Create Payment object
            payment = Payment.objects.create(
                event_registration=EventRegistration.objects.create(user=request.user, event=event),
                amount=current_price,
                order_id=razorpay_order['id'],
                status='pending'
            )

            return JsonResponse({
                'status': 'success',
                'order_id': razorpay_order['id'],
                'amount': order_amount,
                'currency': order_currency,
                'key': settings.RAZORPAY_KEY_ID
            })

        elif action == 'confirm_payment':
            razorpay_payment_id = data.get('razorpay_payment_id')
            razorpay_order_id = data.get('razorpay_order_id')
            razorpay_signature = data.get('razorpay_signature')

            # Verify the payment signature
            try:
                client.utility.verify_payment_signature({
                    'razorpay_order_id': razorpay_order_id,
                    'razorpay_payment_id': razorpay_payment_id,
                    'razorpay_signature': razorpay_signature
                })
            except:
                return JsonResponse({'status': 'error', 'message': 'Invalid payment signature'}, status=400)

            # Update Payment object
            try:
                payment = Payment.objects.get(order_id=razorpay_order_id)
                payment.payment_id = razorpay_payment_id
                payment.status = 'completed'
                payment.save()

                # Update EventRegistration status
                payment.event_registration.status = 'confirmed'
                payment.event_registration.save()
                payment.event_registration.generate_qr_code()

                # Send confirmation email with QR code
                send_registration_confirmation_email(request.user, event, payment.event_registration)

                return JsonResponse({'status': 'success', 'message': 'Payment confirmed'})
            except Payment.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Payment not found'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

def upcoming_events(request):
    current_time = datetime.now()
    user_registrations = EventRegistration.objects.filter(user=request.user)
    registered_event_ids = user_registrations.values_list('event_id', flat=True)
    
    events = Event.objects.filter(
        id__in=registered_event_ids,
        status='upcoming'
    ).order_by('start_datetime')
    registered_events = Event.objects.filter(
        id__in=registered_event_ids,
    ).order_by('start_datetime')
    return render(request, 'events/upcoming_events.html', {'events': events, 'registered_events': registered_events})

@login_required
def event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    registration = EventRegistration.objects.filter(user=request.user, event=event).first()
    is_registered = registration is not None
    is_volunteer = VolunteerRegistration.objects.filter(event=event, user=request.user).exists()

    # Paginate gallery items
    gallery_items = event.gallery_items.all()
    paginator = Paginator(gallery_items, 12)  # Show 12 items per page
    page_number = request.GET.get('page')
    gallery_page = paginator.get_page(page_number)
    
    context = {
        'event': event,
        'registration': registration,
        'is_registered': is_registered,
        'is_volunteer': is_volunteer,
        'gallery_items': gallery_page,
    }
    return render(request, 'events/event_detail.html', context)

@login_required
def event_detail_manage(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.user != event.organizer:
        return redirect('events:event_detail', event_id=event_id)
    
    # Get participants (confirmed registrations)
    participants = EventRegistration.objects.filter(event=event, status='confirmed').select_related('user')
    
    # Get volunteers
    volunteers = VolunteerRegistration.objects.filter(event=event).select_related('user')
    
    # Get payment data
    payments = Payment.objects.filter(event_registration__event=event).select_related('event_registration__user')
    
    # Calculate total payments and amount
    total_payments = payments.count()
    total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'event': event,
        'participants': participants,
        'volunteers': volunteers,
        'payments': payments,
        'total_payments': total_payments,
        'total_amount': total_amount,
    }
    return render(request, 'events/event_detail_manage.html', context)



@login_required
def download_participants_pdf(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    participants = EventRegistration.objects.filter(event=event, status='confirmed').select_related('user')
    
    template_path = 'events/participants_pdf.html'
    context = {'event': event, 'participants': participants}
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="participants_{event.name}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@login_required
def check_attendance(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        updated_count = 0
        for key, value in request.POST.items():
            if key.startswith('attendance_'):
                participant_id = key.split('_')[1]
                try:
                    participant = EventRegistration.objects.get(id=participant_id, event=event)
                    participant.attended = value == 'on'
                    participant.save()
                    updated_count += 1
                except EventRegistration.DoesNotExist:
                    messages.warning(request, f"Participant with ID {participant_id} not found.")
        
        messages.success(request, f"Attendance has been updated successfully. {updated_count} records updated.")
    return redirect('events:event_detail_manage', event_id=event.id)


def send_registration_confirmation_email(user, event, registration):
    subject = f'Registration Confirmation - {event.name}'
    
    # Get the full URL for the QR code
    qr_code_url = None
    if registration.qr_code:
        qr_code_url = registration.qr_code.url
        if not qr_code_url.startswith('http'):
            qr_code_url = settings.SITE_URL + qr_code_url
    
    context = {
        'user': user,
        'event': event,
        'registration': registration,
        'qr_code_url': qr_code_url,
        'site_url': settings.SITE_URL,
        'event_url': f"{settings.SITE_URL}/events/event/{event.id}/"
    }
    
    html_message = render_to_string('events/registration_confirmation_email.html', context)
    plain_message = strip_tags(html_message)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        html_message=html_message,
        fail_silently=False,
    )


@login_required
def download_attendance_sheet(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    participants = EventRegistration.objects.filter(event=event, status='confirmed')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"

    # Add headers
    headers = ['ID', 'Name', 'Email', 'Phone', 'Attended']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Add participant data
    for row, participant in enumerate(participants, start=2):
        ws.cell(row=row, column=1, value=participant.id)
        ws.cell(row=row, column=2, value=f"{participant.user.first_name} {participant.user.last_name}")
        ws.cell(row=row, column=3, value=participant.user.email)
        ws.cell(row=row, column=4, value=participant.user.details.phone_number)
        ws.cell(row=row, column=5, value="")  # Empty cell for marking attendance

    # Generate file name
    file_name = f"attendance_sheet_{slugify(event.name)}.xlsx"

    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
def upload_attendance_sheet(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    if request.method == 'POST' and request.FILES.get('attendance_file'):
        attendance_file = request.FILES['attendance_file']
        
        try:
            wb = openpyxl.load_workbook(attendance_file)
            ws = wb.active

            updated_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 5:
                    continue
                participant_id, _, _, _, attended = row
                if participant_id and attended and str(attended).lower() in ['yes', 'y', 'true', '1']:
                    try:
                        registration = EventRegistration.objects.get(id=participant_id, event=event)
                        registration.attended = True
                        registration.save()
                        updated_count += 1
                    except EventRegistration.DoesNotExist:
                        messages.warning(request, f"Participant with ID {participant_id} not found.")

            messages.success(request, f"Attendance has been updated successfully. {updated_count} records updated.")
        except Exception as e:
            messages.error(request, f"An error occurred while processing the file: {str(e)}")
        
    return redirect('events:event_detail_manage', event_id=event.id)


def predict_event_success_view(request):
    if request.method == 'GET':
        event_category = request.GET.get('event_category')
        community_id = request.GET.get('community_id')
        start_datetime_str = request.GET.get('start_datetime')
        end_datetime_str = request.GET.get('end_datetime')
        description = request.GET.get('description', '')
        is_free = request.GET.get('is_free') == 'true'
        
        community = get_object_or_404(Community, id=community_id)
        
        # Use default values if start or end times are not provided
        now = timezone.now()
        start_datetime = parse_datetime(start_datetime_str) if start_datetime_str else now
        end_datetime = parse_datetime(end_datetime_str) if end_datetime_str else (now + timezone.timedelta(hours=1))
        
        predicted_success = predict_event_success(
            event_category=event_category,
            community_category=community.category.name,
            community_name=community.name,
            start_time=start_datetime,
            end_time=end_datetime,
            description=description,
            is_free=is_free
        )
        
        if predicted_success is None:
            return JsonResponse({'error': 'Unable to make prediction. Please ensure you have generated bulk data and trained the model.'}, status=400)
        
        return JsonResponse({'predicted_success': f'{predicted_success:.2%}'})

@login_required
def update_event_metrics(request, event_id=None):
    try:
        with transaction.atomic():
            if event_id:
                events = Event.objects.filter(id=event_id)
            else:
                events = Event.objects.all()

            for event in events:
                actual_attendees = event.get_actual_attendees()
                event.update_success_metrics(actual_attendees)

        return JsonResponse({'status': 'success', 'message': 'Event metrics updated successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def download_invoice(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    registration = get_object_or_404(EventRegistration, user=request.user, event=event)
    payment = get_object_or_404(Payment, event_registration=registration)

    template_path = 'events/invoice_template.html'
    context = {
        'event': event,
        'registration': registration,
        'payment': payment,
        'user': request.user,
        'website_name': 'EvoCom',
        'website_url': 'https://evocom.com',  # Replace with your actual website URL
    }
    return render(request, template_path, context)

def send_event_reminder_email(event):
    subject = f"Reminder: {event.name} is coming up!"
    participants = EventRegistration.objects.filter(event=event, status='confirmed')
    
    for participant in participants:
        context = {
            'user': participant.user,
            'event': event,
        }
        html_content = render_to_string('events/event_reminder_email.html', context)
        text_content = strip_tags(html_content)
        
        send_mail(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [participant.user.email],
            html_message=html_content,
            fail_silently=True,
        )

@login_required
@require_POST
def register_as_volunteer(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if VolunteerRegistration.objects.filter(user=request.user, event=event).exists():
        return JsonResponse({'status': 'error', 'message': "You have already registered as a volunteer for this event."})
    elif EventRegistration.objects.filter(user=request.user, event=event).exists():
        return JsonResponse({'status': 'error', 'message': "You are already registered as a participant for this event."})
    else:
        VolunteerRegistration.objects.create(user=request.user, event=event)
        return JsonResponse({'status': 'success', 'message': "You have successfully registered as a volunteer."})

@login_required
def manage_volunteers(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.user != event.organizer:
        messages.error(request, "You don't have permission to manage volunteers for this event.")
        return redirect('events:event_detail', event_id=event.id)
    
    volunteers = VolunteerRegistration.objects.filter(event=event)
    
    if request.method == 'POST':
        volunteer_id = request.POST.get('volunteer_id')
        action = request.POST.get('action')
        task = request.POST.get('task')
        
        volunteer = get_object_or_404(VolunteerRegistration, id=volunteer_id)
        
        if action == 'approve':
            volunteer.approved = True
            volunteer.save()
            messages.success(request, f"{volunteer.user.username} has been approved as a volunteer.")
        elif action == 'assign_task':
            volunteer.assigned_task = task
            volunteer.save()
            messages.success(request, f"Task assigned to {volunteer.user.username}.")
        elif action == 'remove':
            volunteer.delete()
            messages.success(request, f"{volunteer.user.username} has been removed from volunteers.")
    
    return render(request, 'events/manage_volunteers.html', {'event': event, 'volunteers': volunteers})

@login_required
def approve_volunteer(request, event_id):
    if request.method == 'POST':
        volunteer_id = request.POST.get('volunteer_id')
        volunteer = get_object_or_404(VolunteerRegistration, id=volunteer_id, event_id=event_id)
        if request.user != volunteer.event.organizer:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
        volunteer.approved = True
        volunteer.save()
        return JsonResponse({'status': 'success', 'message': 'Volunteer approved successfully'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def assign_task(request, event_id):
    if request.method == 'POST':
        volunteer_id = request.POST.get('volunteer_id')
        task = request.POST.get('task')
        volunteer = get_object_or_404(VolunteerRegistration, id=volunteer_id, event_id=event_id)
        if request.user != volunteer.event.organizer:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)
        volunteer.assigned_task = task
        volunteer.save()
        return JsonResponse({'status': 'success', 'message': 'Task assigned successfully'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
@require_POST
def remove_volunteer(request, event_id, volunteer_id):
    try:
        event = Event.objects.get(id=event_id)
        volunteer = VolunteerRegistration.objects.get(id=volunteer_id, event=event)
        
        if request.user == event.organizer or request.user.is_superuser:
            volunteer.delete()
            return JsonResponse({'status': 'success', 'message': 'Volunteer removed successfully.'})
        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Permission denied. Only the event organizer or superuser can remove volunteers.',
                'user_id': request.user.id,
                'is_superuser': request.user.is_superuser,
                'organizer_id': event.organizer.id
            }, status=403)
    except Event.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Event not found.'}, status=404)
    except VolunteerRegistration.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Volunteer registration not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_POST
def email_volunteer(request):
    data = json.loads(request.body)
    email = data.get('email')
    subject = data.get('subject')
    message = data.get('message')
    
    if not all([email, subject, message]):
        return JsonResponse({'status': 'error', 'message': 'Please provide email, subject, and message.'})
    
    try:
        send_mail(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [email],
            fail_silently=False,
        )
        return JsonResponse({'status': 'success', 'message': 'Email sent successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Failed to send email: {str(e)}'})

@login_required
def event_stream(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    is_registered = EventRegistration.objects.filter(event=event, user=request.user).exists()
    is_organizer = event.community.admin == request.user or event.community.eventorganizers_set.filter(user=request.user).exists()
    
    if not (is_registered or is_organizer):
        messages.error(request, "You must be registered for this event to view the stream.")
        return redirect('events:event_detail', event_id=event_id)
    
    context = {
        'event': event,
        'is_organizer': is_organizer,
        'stream_url': event.stream_url,
    }
    return render(request, 'events/event_stream.html', context)

@login_required
def start_stream(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if event.current_status == 'completed':
        return JsonResponse({'error': 'Cannot start stream for completed events'}, status=400)
        
    # Check if user is event organizer or community admin
    is_organizer = (event.organizer == request.user or 
                   event.community.admin == request.user or 
                   event.community.eventorganizerrequest_set.filter(user=request.user, status='approved').exists())
    
    if not is_organizer:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        youtube_url = data.get('youtube_url')
        
        if not youtube_url:
            return JsonResponse({'error': 'YouTube URL is required'}, status=400)
            
        # Validate YouTube URL format
        patterns = [
            r'(?:https?:\/\/)?(?:www\.)?youtube\.com\/watch\?v=[^&]+',
            r'(?:https?:\/\/)?(?:www\.)?youtube\.com\/embed\/[^/?]+',
            r'(?:https?:\/\/)?(?:www\.)?youtu\.be\/[^/?]+',
            r'(?:https?:\/\/)?(?:www\.)?youtube\.com\/live\/[^/?]+'
        ]
        
        if not any(re.match(pattern, youtube_url) for pattern in patterns):
            return JsonResponse({'error': 'Invalid YouTube URL format'}, status=400)
        
        event.youtube_stream_url = youtube_url
        event.is_streaming = True
        event.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Stream started successfully',
            'stream_url': event.youtube_stream_url
        })
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def stop_stream(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Check if user is event organizer or community admin
    is_organizer = (event.organizer == request.user or 
                   event.community.admin == request.user or 
                   event.community.eventorganizerrequest_set.filter(user=request.user, status='approved').exists())
    
    if not is_organizer:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        event.is_streaming = False
        event.youtube_stream_url = None
        event.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Stream stopped successfully'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def stream_status(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    return JsonResponse({
        'is_streaming': event.is_streaming,
        'stream_url': event.stream_url if event.is_streaming else None
    })

@login_required
def generate_stream_key(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if not (event.community.admin == request.user or event.community.eventorganizers_set.filter(user=request.user).exists()):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    stream_key = event.generate_stream_key()
    return JsonResponse({'stream_key': stream_key})

def validate_file(file, item_type):
    # Maximum file size (10MB)
    MAX_SIZE = 10 * 1024 * 1024
    
    if file.size > MAX_SIZE:
        raise ValidationError(f'File size must be no more than {MAX_SIZE/1024/1024}MB')
    
    # Get mime type
    content_type = file.content_type
    
    if item_type in ['image', 'poster']:
        if not content_type.startswith('image/'):
            raise ValidationError('File must be an image')
        # Validate image dimensions
        width, height = get_image_dimensions(file)
        if width > 4096 or height > 4096:
            raise ValidationError('Image dimensions must be no larger than 4096x4096')
            
    elif item_type == 'video':
        allowed_video_types = ['video/mp4', 'video/quicktime', 'video/x-msvideo']
        if content_type not in allowed_video_types:
            raise ValidationError('Video must be in MP4, MOV, or AVI format')
            
    return True

@login_required
def upload_gallery_item(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Check if user is event organizer or community admin
    is_organizer = (event.organizer == request.user or 
                   event.community.admin == request.user or 
                   event.community.eventorganizerrequest_set.filter(user=request.user, status='approved').exists())
    
    if not is_organizer:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            item_type = request.POST.get('item_type')
            file = request.FILES.get('file')
            is_featured = request.POST.get('is_featured') == 'true'
            
            if not all([title, item_type, file]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # Validate file
            try:
                validate_file(file, item_type)
            except ValidationError as e:
                return JsonResponse({'error': str(e)}, status=400)
            
            gallery_item = EventGalleryItem.objects.create(
                event=event,
                title=title,
                description=description,
                file=file,
                item_type=item_type,
                is_featured=is_featured
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Gallery item uploaded successfully',
                'item': {
                    'id': gallery_item.id,
                    'title': gallery_item.title,
                    'url': gallery_item.file.url
                }
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
def delete_gallery_item(request, event_id, item_id):
    gallery_item = get_object_or_404(EventGalleryItem, id=item_id, event_id=event_id)
    
    # Check if user is event organizer or community admin
    is_organizer = (gallery_item.event.organizer == request.user or 
                   gallery_item.event.community.admin == request.user or 
                   gallery_item.event.community.eventorganizerrequest_set.filter(user=request.user, status='approved').exists())
    
    if not is_organizer:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            # Delete the file from storage
            gallery_item.file.delete(save=False)
            # Delete the database record
            gallery_item.delete()
            return JsonResponse({
                'success': True,
                'message': 'Gallery item deleted successfully'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
def event_organizer_dashboard(request):
    # ... existing code ...
    
    # Get unread messages count
    unread_messages_count = ChatMessage.objects.filter(
        receiver=request.user,
        is_read=False
    ).count()
    
    context = {
        # ... existing context ...
        'unread_messages_count': unread_messages_count,
    }
    return render(request, 'events/event_organizer_dashboard.html', context)

def event_detail_api(request, event_id):
    try:
        event = Event.objects.get(id=event_id)
        # Get registration count from related registrations
        registration_count = event.eventregistration_set.count()
        
        data = {
            'name': event.name,
            'description': event.description,
            'start_datetime': event.start_datetime.isoformat(),
            'end_datetime': event.end_datetime.isoformat(),
            'location': event.location,
            'registration_count': registration_count,
            'registration_fee': float(event.registration_fee),
            'cover_image': event.cover_image.url if event.cover_image else None,
        }
        return JsonResponse(data)
    except Event.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)

def generate_receipt_pdf(event, user, payment):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Set the font to the one that supports the rupee symbol
    p.setFont("DejaVuSans", 12)

    # Add content to the PDF
    p.drawString(50, height - 50, f"Total Revenue: ₹{payment.amount}")

    # Continue with the rest of your PDF generation logic...

@login_required
def initiate_payment(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Use current_price instead of registration_fee
    amount = event.current_price
    
    # Create Razorpay order
    order_amount = int(float(amount) * 100)  # Convert to paise
    order_currency = 'INR'
    
    # ... rest of the payment processing logic ...

@login_required
def view_ticket(request, registration_id):
    registration = get_object_or_404(EventRegistration, id=registration_id, user=request.user)
    event = registration.event
    
    # Generate QR code if it doesn't exist
    if not registration.qr_code:
        registration.generate_qr_code()
        registration.save()
    
    context = {
        'user': request.user,
        'event': event,
        'registration': registration,
    }
    
    return render(request, 'events/ticket.html', context)

@login_required
def download_ticket(request, registration_id):
    registration = get_object_or_404(EventRegistration, id=registration_id, user=request.user)
    event = registration.event
    
    # Create a BytesIO buffer for the PDF
    buffer = BytesIO()
    
    # Create the PDF object using reportlab
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Add content to PDF
    p.setFont("Helvetica-Bold", 24)
    p.drawString(50, height - 50, event.name)
    
    # Add QR Code
    if registration.qr_code:
        p.drawImage(registration.qr_code.path, 50, height - 300, width=200, height=200)
    
    # Add ticket details
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 350, f"Ticket ID: {registration.ticket_id}")
    p.drawString(50, height - 370, f"Attendee: {request.user.get_full_name()}")
    p.drawString(50, height - 390, f"Event Date: {event.start_datetime.strftime('%B %d, %Y, %I:%M %p')}")
    p.drawString(50, height - 410, f"Venue: {event.location}")
    
    p.showPage()
    p.save()
    
    # Get the value of the BytesIO buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create the HTTP response with PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{registration.ticket_id}.pdf"'
    response.write(pdf)
    
    return response

@login_required
def scan_qr_code(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Check if user is event organizer
    if event.organizer != request.user:
        messages.error(request, "You don't have permission to manage attendance.")
        return redirect('events:event_detail', event_id=event_id)
    
    context = {
        'event': event,
    }
    return render(request, 'events/scan_qr.html', context)

@login_required
@require_POST
def verify_ticket(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Check if user is event organizer
    if event.organizer != request.user:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    ticket_id = request.POST.get('ticket_id')
    try:
        registration = EventRegistration.objects.get(ticket_id=ticket_id, event=event)
        if registration.attended:
            return JsonResponse({
                'success': False,
                'message': 'Attendance already marked for this ticket'
            })
        
        registration.attended = True
        registration.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Attendance marked successfully',
            'attendee': registration.user.get_full_name()
        })
    except EventRegistration.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Invalid ticket'
        })
